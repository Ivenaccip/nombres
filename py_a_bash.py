import openpyxl

#Abrir y leer Excel

BD_influencers = openpyxl.load_workbook('Influencers_links.xlsx')
sheet = BD_influencers.active

# Abrir un nuevo sheet

archivo = open("C:/Users/Ivenaccip/Documents/Kroon/archivo.txt", "w")

#Encontrar patrón de la info

for i in range(2, 128):
    nombre_inf = sheet.cell(row = i, column = 1).value
    insta = sheet.cell(row = i, column = 2).value

#Pasar todo a una .txt

    archivo.write(f"[{nombre_inf}]='{insta}' \n")

archivo.close()
